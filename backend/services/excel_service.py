import pandas as pd
import logging
from typing import List, Dict, Tuple
from pathlib import Path
from datetime import datetime
from models import models
import os

logger = logging.getLogger(__name__)

class ExcelService:
    def __init__(self):
        self.supported_extensions = ['.xlsx', '.xls']
        self.exports_dir = Path("exports")
        self.exports_dir.mkdir(exist_ok=True)
    
    def read_boq_file(self, file_path: str) -> List[Dict]:
        """Read BOQ items from Excel file"""
        try:
            df = pd.read_excel(file_path, sheet_name=0)
            
            # Clean column names by stripping whitespace
            df.columns = df.columns.str.strip()
            # print("_____Cleaned column names_____", list(df.columns))
            
            items = []
            
            for _, row in df.iterrows():
                # print("_____row_____", row)
                # print("_____row types_____", row.dtypes)
                # print("_____Serial Number raw_____", row.get('Serial Number'), "Type:", type(row.get('Serial Number')))
                # print("_____Description raw_____", row.get('Description'), "Type:", type(row.get('Description')))
                
                # Check if we have the minimum required data
                if pd.notna(row.get('Section Number', '')) and pd.notna(row.get('Description', '')):
                    section_number = str(row.get('Section Number', '')).strip()
                    parts = section_number.split('.')
                    subsection = '.'.join(parts[:3]) if len(parts) >= 3 else section_number
                    
                    # Extract all values with proper null handling
                    serial_number = row.get('Serial Number')
                    structure = row.get('Structure')
                    system = row.get('System')
                    description = row.get('Description')
                    unit = row.get('Unit')
                    original_contract_quantity = row.get('Original Contract Quantity')
                    price = row.get('Price')
                    total_contract_sum = row.get('Total Contract Sum')
                    estimated_quantity = row.get('Estimated Quantity')
                    quantity_submitted = row.get('Quantity Submitted')
                    internal_quantity = row.get('Internal Quantity')
                    approved_by_project_manager = row.get('Approved by Project Manager')
                    total_estimate = row.get('Total Estimate')
                    total_submitted = row.get('Total Submitted')
                    internal_total = row.get('Internal Total')
                    total_approved_by_project_manager = row.get('Total Approved by Project Manager')
                    notes = row.get('NOTES')
                    
                    # print("_____Extracted values_____")
                    # print("serial_number:", serial_number, "pd.notna:", pd.notna(serial_number))
                    # print("description:", description, "pd.notna:", pd.notna(description))
                    
                    # More robust null checking
                    def safe_float(value):
                        if pd.isna(value) or value == '' or value is None:
                            return None
                        try:
                            return float(value)
                        except (ValueError, TypeError):
                            return None
                    
                    def safe_int(value):
                        if pd.isna(value) or value == '' or value is None:
                            return None
                        try:
                            return int(float(value))  # Convert float to int for cases like 59.0 -> 59
                        except (ValueError, TypeError):
                            return None
                    
                    def safe_str(value):
                        if pd.isna(value) or value is None:
                            return ''
                        return str(value).strip()
                    
                    item = {
                        'serial_number': safe_int(serial_number),
                        'structure': safe_int(structure),
                        'system': safe_str(system),
                        'section_number': section_number,
                        'description': safe_str(description),
                        'unit': safe_str(unit),
                        'original_contract_quantity': safe_float(original_contract_quantity) or 0,
                        'price': safe_float(price) or 0,
                        'total_contract_sum': safe_float(total_contract_sum) or 0,
                        'estimated_quantity': safe_float(estimated_quantity) or 0,
                        'quantity_submitted': safe_float(quantity_submitted) or 0,
                        'internal_quantity': safe_float(internal_quantity) or 0,
                        'approved_by_project_manager': safe_float(approved_by_project_manager) or 0,
                        'total_estimate': safe_float(total_estimate) or 0,
                        'total_submitted': safe_float(total_submitted) or 0,
                        'internal_total': safe_float(internal_total) or 0,
                        'total_approved_by_project_manager': safe_float(total_approved_by_project_manager) or 0,
                        'approved_signed_quantity': 0.0,  # Default value for new column
                        'approved_signed_total': 0.0,  # Default value for new column
                        'notes': safe_str(notes),
                        'subsection': subsection
                    }
                    # print("_____item_____", item)
                    items.append(item)
            
            logger.info(f"Successfully read {len(items)} items from BOQ file")
            return items
            
        except Exception as e:
            logger.error(f"Error reading BOQ file: {str(e)}")
            raise
    

    def read_calculation_sheet_data(self, file_path: str) -> Dict:
        """
        Read calculation sheet data from Excel file based on specific cell locations
        Returns: {
            'calculation_sheet_no': str,
            'drawing_no': str,
            'description': str,
            'entries': [
                {
                    'section_number': str,
                    'estimated_quantity': float,
                    'quantity_submitted': float
                }
            ]
        }
        """
        try:
            print("_____ok_____")
            # Read Excel file without headers to access specific cells

            # if os.path.exists(file_path):
            #     print("_____file exists_____")
            # else:
            #     print("_____file does not exist_____")

            df = pd.read_excel(file_path, sheet_name="Calculation", header=None)
            df.to_csv("output.csv", index=False, header=False)
            # print(df)
            
            # Extract header information from specific cells
            calculation_sheet_no = str(df.iloc[0, 2]).strip() if pd.notna(df.iloc[0, 2]) else ""  # C1
            drawing_no = str(df.iloc[1, 2]).strip() if pd.notna(df.iloc[1, 2]) else ""  # C2
            description = str(df.iloc[2, 2]).strip() if pd.notna(df.iloc[2, 2]) else ""  # C3
            
            # Validate required fields
            if not calculation_sheet_no or not drawing_no or not description:
                raise ValueError(f"Missing required header information in {file_path}")
            
            entries = []
            
            col_index = 4
            
            while col_index < df.shape[1]:
                # Check if section number exists in row 4 (J5, K5, L5, etc.)
                section_number = df.iloc[4, col_index]
                
                if pd.notna(section_number) and str(section_number).strip():
                    section_number = str(section_number).strip()
                    
                    # Get estimated quantity from row 5 (J6, K6, L6, etc.)
                    estimated_quantity = df.iloc[5, col_index]
                    estimated_quantity = float(estimated_quantity) if pd.notna(estimated_quantity) else 0.0
                    
                    # Get quantity submitted from row 23 (J24, K24, L24, etc.)
                    quantity_submitted = df.iloc[23, col_index]
                    quantity_submitted = float(quantity_submitted) if pd.notna(quantity_submitted) else 0.0
                    
                    entry = {
                        'section_number': section_number,
                        'estimated_quantity': estimated_quantity,
                        'quantity_submitted': quantity_submitted
                    }
                    entries.append(entry)
                
                col_index += 1
            
            if not entries:
                logger.warning(f"No valid entries found in {file_path}")
            
            result = {
                'calculation_sheet_no': calculation_sheet_no,
                'drawing_no': drawing_no,
                'description': description,
                'entries': entries
            }
            
            logger.info(f"Successfully read calculation sheet data from {file_path}: {len(entries)} entries")
            return result
            
        except Exception as e:
            logger.error(f"Error reading calculation sheet data from {file_path}: {str(e)}")
            raise
    
    def process_excel_file(self, file_path: Path) -> Tuple[List[Dict], List[str]]:
        """
        Process an Excel file and return items and errors
        Returns: (items, errors)
        """
        items = []
        errors = []
        
        try:
            # Try to process as BOQ file first
            try:
                items = self.read_boq_file(str(file_path))
                logger.info(f"Successfully processed {len(items)} BOQ items from {file_path.name}")
                return items, errors
            except Exception as e:
                logger.warning(f"Failed to process {file_path.name} as BOQ file: {str(e)}")
                errors.append(f"Failed to process as BOQ file: {str(e)}")
            
            # If BOQ processing failed, try as calculation file
            try:
                calculation_entries = self.read_calculation_file(str(file_path))
                # For calculation files, we need to create BOQ items from the entries
                # This is a simplified approach - you might need to adjust based on your needs
                if calculation_entries:
                    # Create a single BOQ item from the calculation file
                    item = {
                        'serial_number': None,
                        'structure': None,
                        'section_number': f"CALC_{file_path.stem}",
                        'description': f"Calculation entries from {file_path.name}",
                        'unit': 'N/A',
                        'original_contract_quantity': 0,
                        'price': 0,
                        'total_contract_sum': 0,
                        'estimated_quantity': sum(entry.get('estimated_quantity', 0) for entry in calculation_entries),
                        'quantity_submitted': sum(entry.get('quantity_submitted', 0) for entry in calculation_entries),
                        'internal_quantity': sum(entry.get('internal_quantity', 0) for entry in calculation_entries),
                        'approved_by_project_manager': sum(entry.get('approved_by_project_manager', 0) for entry in calculation_entries),
                        'total_estimate': sum(entry.get('estimated_quantity', 0) for entry in calculation_entries),
                        'total_submitted': sum(entry.get('quantity_submitted', 0) for entry in calculation_entries),
                        'internal_total': sum(entry.get('internal_quantity', 0) for entry in calculation_entries),
                        'total_approved_by_project_manager': sum(entry.get('approved_by_project_manager', 0) for entry in calculation_entries),
                        'notes': f"Processed from calculation file: {file_path.name}",
                        'subsection': 'CALCULATIONS'
                    }
                    items = [item]
                    logger.info(f"Successfully processed calculation file {file_path.name} into {len(items)} BOQ items")
                    return items, errors
            except Exception as e:
                logger.warning(f"Failed to process {file_path.name} as calculation file: {str(e)}")
                errors.append(f"Failed to process as calculation file: {str(e)}")
            
            # If both failed, return empty items with errors
            if not items:
                errors.append(f"Could not process {file_path.name} as either BOQ or calculation file")
            
            return items, errors
            
        except Exception as e:
            error_msg = f"Unexpected error processing {file_path.name}: {str(e)}"
            logger.error(error_msg)
            errors.append(error_msg)
            return items, errors 

    def export_single_concentration_sheet(self, sheet, boq_item, entries):
        """Export a single concentration sheet to Excel with specific format: 3 tables"""
        try:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"concentration_sheet_{sheet.id}_{timestamp}.xlsx"
            filepath = self.exports_dir / filename
            
            # Create Excel writer
            with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                
                # Single sheet with all content
                sheet_name = 'Concentration Sheet'
                current_row = 0
                
                # First Table: Project Information (2 rows, 4 columns)
                project_headers = ['Contract No', 'Developer Name', 'Project Name', 'Contractor in Charge']
                project_values = [
                    sheet.contract_no or 'N/A',
                    sheet.developer_name or 'N/A', 
                    sheet.project_name or 'N/A',
                    sheet.contractor_in_charge or 'N/A'
                ]
                
                project_data = [project_headers, project_values]
                df_project = pd.DataFrame(project_data)
                df_project.to_excel(writer, sheet_name=sheet_name, index=False, header=False, startrow=current_row)
                current_row += 3  # 2 rows + 1 spacing row
                
                # Second Table: BOQ Item Details (2 rows, 5 columns)
                boq_headers = ['Section No', 'Contract Quantity', 'Unit', 'Price', 'Description']
                boq_values = [
                    boq_item.section_number,
                    f"{boq_item.original_contract_quantity:,.2f}",
                    boq_item.unit,
                    f"{boq_item.price:,.2f}",
                    boq_item.description or ''
                ]
                
                boq_data = [boq_headers, boq_values]
                df_boq = pd.DataFrame(boq_data)
                df_boq.to_excel(writer, sheet_name=sheet_name, index=False, header=False, startrow=current_row)
                current_row += 3  # 2 rows + 1 spacing row
                
                # Third Table: Concentration Entries (following the order shown on concentration sheets page)
                if entries:
                    # Column order as shown on concentration sheets page
                    entries_headers = ['Description', 'Calculation Sheet No', 'Drawing No', 'Estimated Quantity', 
                                     'Quantity Submitted', 'Internal Quantity', 'Approved by Project Manager', 'Notes']
                    
                    entries_data = [entries_headers]
                    for entry in entries:
                        entries_data.append([
                            entry.description or '',
                            entry.calculation_sheet_no or '',
                            entry.drawing_no or '',
                            f"{entry.estimated_quantity:,.2f}",
                            f"{entry.quantity_submitted:,.2f}",
                            f"{entry.internal_quantity:,.2f}",
                            f"{entry.approved_by_project_manager:,.2f}",
                            entry.notes or ''
                        ])
                    
                    # Add totals row
                    total_estimate = sum(entry.estimated_quantity for entry in entries)
                    total_submitted = sum(entry.quantity_submitted for entry in entries)
                    total_internal = sum(entry.internal_quantity for entry in entries)
                    total_approved = sum(entry.approved_by_project_manager for entry in entries)
                    
                    entries_data.append([
                        'TOTALS',
                        '',
                        '',
                        f"{total_estimate:,.2f}",
                        f"{total_submitted:,.2f}",
                        f"{total_internal:,.2f}",
                        f"{total_approved:,.2f}",
                        ''
                    ])
                    
                    df_entries = pd.DataFrame(entries_data)
                    df_entries.to_excel(writer, sheet_name=sheet_name, index=False, header=False, startrow=current_row)
                
                # Apply formatting to the single sheet
                workbook = writer.book
                worksheet = workbook[sheet_name]
                
                # Auto-adjust column widths
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
                
                # Style header rows and apply right alignment for RTL
                from openpyxl.styles import Font, PatternFill, Alignment
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                header_alignment = Alignment(horizontal="right", vertical="center")
                data_alignment = Alignment(horizontal="right", vertical="center")
                
                # Apply right alignment to all cells for RTL layout
                for row in worksheet.iter_rows():
                    for cell in row:
                        cell.alignment = data_alignment
                
                # Style first table header row (Project Information - row 1)
                for cell in worksheet[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                
                # Style second table header row (BOQ Item Details - row 4)
                for cell in worksheet[4]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                
                # Style third table header row (Concentration Entries - row 7)
                if entries:
                    entries_header_row = 7  # After project info (3 rows) + spacing (1 row) + boq details (3 rows)
                    for cell in worksheet[entries_header_row]:
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = header_alignment
                    
                    # Style totals row (last row of entries table)
                    totals_row = entries_header_row + len(entries_data)
                    for cell in worksheet[totals_row]:
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill(start_color="87CEEB", end_color="87CEEB", fill_type="solid")
                
            
            logger.info(f"Generated concentration sheet Excel with single sheet RTL layout: {filepath}")
            return str(filepath)
            
        except Exception as e:
            logger.error(f"Error generating concentration sheet Excel: {str(e)}")
            raise

    def export_all_concentration_sheets(self, sheets, db_session):
        """Export all concentration sheets to Excel with separate sheets (sheet name = section number)"""
        try:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"all_concentration_sheets_{timestamp}.xlsx"
            filepath = self.exports_dir / filename
            
            # Create Excel writer
            with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                
                # Create individual sheets for each concentration sheet
                for sheet in sheets:
                    # Get the associated BOQ item
                    boq_item = db_session.query(models.BOQItem).filter(
                        models.BOQItem.id == sheet.boq_item_id
                    ).first()
                    
                    if not boq_item:
                        continue
                    
                    # Get all entries for this concentration sheet
                    entries = db_session.query(models.ConcentrationEntry).filter(
                        models.ConcentrationEntry.concentration_sheet_id == sheet.id
                    ).order_by(models.ConcentrationEntry.id).all()
                    
                    # Use section number as sheet name (Excel sheet name limit is 31 characters)
                    sheet_name = str(boq_item.section_number)[:31]
                    
                    current_row = 0
                    
                    # First Table: Project Information (2 rows, 4 columns)
                    project_headers = ['Contract No', 'Developer Name', 'Project Name', 'Contractor in Charge']
                    project_values = [
                            sheet.contract_no or 'N/A',
                            sheet.developer_name or 'N/A',
                        sheet.project_name or 'N/A',
                        sheet.contractor_in_charge or 'N/A'
                    ]
                    
                    project_data = [project_headers, project_values]
                    df_project = pd.DataFrame(project_data)
                    df_project.to_excel(writer, sheet_name=sheet_name, index=False, header=False, startrow=current_row)
                    current_row += 3  # 2 rows + 1 spacing row
                    
                    # Second Table: BOQ Item Details (2 rows, 5 columns)
                    boq_headers = ['Section No', 'Contract Quantity', 'Unit', 'Price', 'Description']
                    boq_values = [
                            boq_item.section_number,
                            f"{boq_item.original_contract_quantity:,.2f}",
                        boq_item.unit,
                        f"{boq_item.price:,.2f}",
                        boq_item.description or ''
                    ]
                    
                    boq_data = [boq_headers, boq_values]
                    df_boq = pd.DataFrame(boq_data)
                    df_boq.to_excel(writer, sheet_name=sheet_name, index=False, header=False, startrow=current_row)
                    current_row += 3  # 2 rows + 1 spacing row
                    
                    # Third Table: Concentration Entries (following the order shown on concentration sheets page)
                    if entries:
                        # Column order as shown on concentration sheets page
                        entries_headers = ['Description', 'Calculation Sheet No', 'Drawing No', 'Estimated Quantity', 
                                         'Quantity Submitted', 'Internal Quantity', 'Approved by Project Manager', 'Notes']
                        
                        entries_data = [entries_headers]
                        for entry in entries:
                            entries_data.append([
                                entry.description or '',
                                entry.calculation_sheet_no or '',
                                entry.drawing_no or '',
                                f"{entry.estimated_quantity:,.2f}",
                                f"{entry.quantity_submitted:,.2f}",
                                f"{entry.internal_quantity:,.2f}",
                                f"{entry.approved_by_project_manager:,.2f}",
                                entry.notes or ''
                            ])
                        
                        # Add totals row
                        total_estimate = sum(entry.estimated_quantity for entry in entries)
                        total_submitted = sum(entry.quantity_submitted for entry in entries)
                        total_internal = sum(entry.internal_quantity for entry in entries)
                        total_approved = sum(entry.approved_by_project_manager for entry in entries)
                        
                        entries_data.append([
                            'TOTALS',
                            '',
                            '',
                            f"{total_estimate:,.2f}",
                            f"{total_submitted:,.2f}",
                            f"{total_internal:,.2f}",
                            f"{total_approved:,.2f}",
                            ''
                        ])
                        
                        df_entries = pd.DataFrame(entries_data)
                        df_entries.to_excel(writer, sheet_name=sheet_name, index=False, header=False, startrow=current_row)
                    
                    # Apply formatting to the sheet
                    workbook = writer.book
                    worksheet = workbook[sheet_name]
                    
                    # Auto-adjust column widths
                    for column in worksheet.columns:
                        max_length = 0
                        column_letter = column[0].column_letter
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        adjusted_width = min(max_length + 2, 50)
                        worksheet.column_dimensions[column_letter].width = adjusted_width
                    
                    # Style header rows and apply right alignment for RTL
                    from openpyxl.styles import Font, PatternFill, Alignment
                    header_font = Font(bold=True, color="FFFFFF")
                    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                    header_alignment = Alignment(horizontal="right", vertical="center")
                    data_alignment = Alignment(horizontal="right", vertical="center")
                    
                    # Apply right alignment to all cells for RTL layout
                    for row in worksheet.iter_rows():
                        for cell in row:
                            cell.alignment = data_alignment
                    
                    # Style first table header row (Project Information - row 1)
                    for cell in worksheet[1]:
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = header_alignment
                    
                    # Style second table header row (BOQ Item Details - row 4)
                    for cell in worksheet[4]:
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = header_alignment
                    
                    # Style third table header row (Concentration Entries - row 7)
                    if entries:
                        entries_header_row = 7  # After project info (3 rows) + spacing (1 row) + boq details (3 rows)
                        for cell in worksheet[entries_header_row]:
                            cell.font = header_font
                            cell.fill = header_fill
                            cell.alignment = header_alignment
                        
                        # Style totals row (last row of entries table)
                        totals_row = entries_header_row + len(entries_data)
                        for cell in worksheet[totals_row]:
                            cell.font = Font(bold=True)
                            cell.fill = PatternFill(start_color="87CEEB", end_color="87CEEB", fill_type="solid")
                
                logger.info(f"Generated all concentration sheets Excel with {len(sheets)} sheets: {filepath}")
            return str(filepath)
            
        except Exception as e:
            logger.error(f"Error generating all concentration sheets Excel: {str(e)}")
            raise

    def export_structures_summary(self, summaries):
        """Export structures summary to Excel"""
        try:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"structures_summary_{timestamp}.xlsx"
            filepath = self.exports_dir / filename
            
            if not summaries:
                raise ValueError("No data to export")
            
            # Convert to DataFrame
            df = pd.DataFrame(summaries)
            
            # Create a copy of original numeric data for totals calculation
            original_numeric_data = {}
            for col in df.columns:
                if df[col].dtype in ['float64', 'int64']:
                    original_numeric_data[col] = df[col].copy()
                elif col in ['total_contract_sum', 'total_estimate', 'total_submitted', 'internal_total', 'total_approved', 'approved_signed_total'] or col.startswith('total_updated_contract_sum_'):
                    # Force conversion to numeric for known total columns
                    try:
                        numeric_series = pd.to_numeric(df[col], errors='coerce')
                        original_numeric_data[col] = numeric_series.fillna(0)
                    except:
                        pass
            
            # Format only price and sum/total columns (not quantity columns)
            for col in df.columns:
                if col in df.columns and df[col].dtype in ['float64', 'int64']:
                    if ('total' in col.lower() or 'estimate' in col.lower() or 'submitted' in col.lower() or 'approved' in col.lower()) and 'quantity' not in col.lower():
                        df[col] = df[col].apply(lambda x: f"${x:,.2f}" if pd.notna(x) and isinstance(x, (int, float)) else "$0.00")
            
            # Define columns that should have grand totals (same as BOQ items export)
            total_columns = {
                'total_contract_sum',
                'total_estimate',
                'total_submitted',
                'internal_total',
                'total_approved',
                'approved_signed_total'
            }
            # Add updated contract sum columns
            for col in df.columns:
                if col.startswith('total_updated_contract_sum_'):
                    total_columns.add(col)
            
            # Calculate grand totals row using original numeric data - only for specified columns
            totals_row = {}
            for col in df.columns:
                if col == 'structure':
                    totals_row[col] = "GRAND TOTAL"
                elif col in total_columns:
                    if col in original_numeric_data:
                        # Calculate totals from original numeric data
                        total_value = sum([val for val in original_numeric_data[col] if pd.notna(val)])
                        # Apply currency formatting for all total columns
                        totals_row[col] = f"${total_value:,.2f}" if isinstance(total_value, (int, float)) else "$0.00"
                    else:
                        # If column not in original_numeric_data, try to extract numeric values from formatted data
                        try:
                            # Extract numeric values from the formatted currency strings in the DataFrame
                            numeric_values = []
                            for val in df[col]:
                                if pd.notna(val) and isinstance(val, str) and val.startswith('$'):
                                    # Remove $ and commas, convert to float
                                    clean_val = val.replace('$', '').replace(',', '')
                                    try:
                                        numeric_values.append(float(clean_val))
                                    except:
                                        pass
                                elif pd.notna(val) and isinstance(val, (int, float)):
                                    numeric_values.append(val)
                            total_value = sum(numeric_values) if numeric_values else 0
                            totals_row[col] = f"${total_value:,.2f}"
                        except:
                            totals_row[col] = "$0.00"
                else:
                    totals_row[col] = ""
            
            df_totals = pd.DataFrame([totals_row])
            df_final = pd.concat([df, df_totals], ignore_index=True)
            
            # Export to Excel
            with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                df_final.to_excel(writer, sheet_name='Structures Summary', index=False)
                
                # Apply formatting
                workbook = writer.book
                worksheet = writer.sheets['Structures Summary']
                
                # Auto-adjust column widths
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
                
                # Style header row
                from openpyxl.styles import Font, PatternFill, Alignment
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                header_alignment = Alignment(horizontal="center", vertical="center")
                
                for cell in worksheet[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                
                # Style totals row
                totals_row_num = len(df_final) + 1
                totals_font = Font(bold=True, color="FFFFFF")
                totals_fill = PatternFill(start_color="C0504D", end_color="C0504D", fill_type="solid")
                totals_alignment = Alignment(horizontal="center", vertical="center")
                
                for col in range(1, worksheet.max_column + 1):
                    cell = worksheet.cell(row=totals_row_num, column=col)
                    cell.font = totals_font
                    cell.fill = totals_fill
                    cell.alignment = totals_alignment
            
            logger.info(f"Generated structures summary Excel: {filepath}")
            return str(filepath)
            
        except Exception as e:
            logger.error(f"Error generating structures summary Excel: {str(e)}")
            raise

    def export_systems_summary(self, summaries):
        """Export systems summary to Excel"""
        try:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"systems_summary_{timestamp}.xlsx"
            filepath = self.exports_dir / filename
            
            if not summaries:
                raise ValueError("No data to export")
            
            # Convert to DataFrame
            df = pd.DataFrame(summaries)
            
            # Create a copy of original numeric data for totals calculation
            original_numeric_data = {}
            for col in df.columns:
                if df[col].dtype in ['float64', 'int64']:
                    original_numeric_data[col] = df[col].copy()
                elif col in ['total_contract_sum', 'total_estimate', 'total_submitted', 'internal_total', 'total_approved', 'approved_signed_total'] or col.startswith('total_updated_contract_sum_'):
                    # Force conversion to numeric for known total columns
                    try:
                        numeric_series = pd.to_numeric(df[col], errors='coerce')
                        original_numeric_data[col] = numeric_series.fillna(0)
                    except:
                        pass
            
            # Format only price and sum/total columns (not quantity columns)
            for col in df.columns:
                if col in df.columns and df[col].dtype in ['float64', 'int64']:
                    if ('total' in col.lower() or 'estimate' in col.lower() or 'submitted' in col.lower() or 'approved' in col.lower()) and 'quantity' not in col.lower():
                        df[col] = df[col].apply(lambda x: f"${x:,.2f}" if pd.notna(x) and isinstance(x, (int, float)) else "$0.00")
            
            # Define columns that should have grand totals (same as BOQ items export)
            total_columns = {
                'total_contract_sum',
                'total_estimate',
                'total_submitted',
                'internal_total',
                'total_approved',
                'approved_signed_total'
            }
            # Add updated contract sum columns
            for col in df.columns:
                if col.startswith('total_updated_contract_sum_'):
                    total_columns.add(col)
            
            # Calculate grand totals row using original numeric data - only for specified columns
            totals_row = {}
            for col in df.columns:
                if col == 'system':
                    totals_row[col] = "GRAND TOTAL"
                elif col in total_columns:
                    if col in original_numeric_data:
                        # Calculate totals from original numeric data
                        total_value = sum([val for val in original_numeric_data[col] if pd.notna(val)])
                        # Apply currency formatting for all total columns
                        totals_row[col] = f"${total_value:,.2f}" if isinstance(total_value, (int, float)) else "$0.00"
                    else:
                        # If column not in original_numeric_data, try to extract numeric values from formatted data
                        try:
                            # Extract numeric values from the formatted currency strings in the DataFrame
                            numeric_values = []
                            for val in df[col]:
                                if pd.notna(val) and isinstance(val, str) and val.startswith('$'):
                                    # Remove $ and commas, convert to float
                                    clean_val = val.replace('$', '').replace(',', '')
                                    try:
                                        numeric_values.append(float(clean_val))
                                    except:
                                        pass
                                elif pd.notna(val) and isinstance(val, (int, float)):
                                    numeric_values.append(val)
                            total_value = sum(numeric_values) if numeric_values else 0
                            totals_row[col] = f"${total_value:,.2f}"
                        except:
                            totals_row[col] = "$0.00"
                else:
                    totals_row[col] = ""
            
            df_totals = pd.DataFrame([totals_row])
            df_final = pd.concat([df, df_totals], ignore_index=True)
            
            # Export to Excel
            with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                df_final.to_excel(writer, sheet_name='Systems Summary', index=False)
                
                # Apply formatting
                workbook = writer.book
                worksheet = writer.sheets['Systems Summary']
                
                # Auto-adjust column widths
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
                
                # Style header row
                from openpyxl.styles import Font, PatternFill, Alignment
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                header_alignment = Alignment(horizontal="center", vertical="center")
                
                for cell in worksheet[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                
                # Style totals row
                totals_row_num = len(df_final) + 1
                totals_font = Font(bold=True, color="FFFFFF")
                totals_fill = PatternFill(start_color="C0504D", end_color="C0504D", fill_type="solid")
                totals_alignment = Alignment(horizontal="center", vertical="center")
                
                for col in range(1, worksheet.max_column + 1):
                    cell = worksheet.cell(row=totals_row_num, column=col)
                    cell.font = totals_font
                    cell.fill = totals_fill
                    cell.alignment = totals_alignment
            
            logger.info(f"Generated systems summary Excel: {filepath}")
            return str(filepath)
            
        except Exception as e:
            logger.error(f"Error generating systems summary Excel: {str(e)}")
            raise

    def export_subsections_summary(self, summaries):
        """Export subsections summary to Excel"""
        try:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"subsections_summary_{timestamp}.xlsx"
            filepath = self.exports_dir / filename
            
            if not summaries:
                raise ValueError("No data to export")
            
            # Convert to DataFrame
            df = pd.DataFrame(summaries)
            
            # Create a copy of original numeric data for totals calculation
            original_numeric_data = {}
            for col in df.columns:
                if df[col].dtype in ['float64', 'int64']:
                    original_numeric_data[col] = df[col].copy()
                elif col in ['total_contract_sum', 'total_estimate', 'total_submitted', 'internal_total', 'total_approved', 'approved_signed_total'] or col.startswith('total_updated_contract_sum_'):
                    # Force conversion to numeric for known total columns
                    try:
                        numeric_series = pd.to_numeric(df[col], errors='coerce')
                        original_numeric_data[col] = numeric_series.fillna(0)
                    except:
                        pass
            
            # Format only price and sum/total columns (not quantity columns)
            for col in df.columns:
                if col in df.columns and df[col].dtype in ['float64', 'int64']:
                    if ('total' in col.lower() or 'estimate' in col.lower() or 'submitted' in col.lower() or 'approved' in col.lower()) and 'quantity' not in col.lower():
                        df[col] = df[col].apply(lambda x: f"${x:,.2f}" if pd.notna(x) and isinstance(x, (int, float)) else "$0.00")
            
            # Define columns that should have grand totals (same as BOQ items export)
            total_columns = {
                'total_contract_sum',
                'total_estimate',
                'total_submitted',
                'internal_total',
                'total_approved',
                'approved_signed_total'
            }
            # Add updated contract sum columns
            for col in df.columns:
                if col.startswith('total_updated_contract_sum_'):
                    total_columns.add(col)
            
            # Calculate grand totals row using original numeric data - only for specified columns
            totals_row = {}
            for col in df.columns:
                if col == 'subsection':
                    totals_row[col] = "GRAND TOTAL"
                elif col in total_columns:
                    if col in original_numeric_data:
                        # Calculate totals from original numeric data
                        total_value = sum([val for val in original_numeric_data[col] if pd.notna(val)])
                        # Apply currency formatting for all total columns
                        totals_row[col] = f"${total_value:,.2f}" if isinstance(total_value, (int, float)) else "$0.00"
                    else:
                        # If column not in original_numeric_data, try to extract numeric values from formatted data
                        try:
                            # Extract numeric values from the formatted currency strings in the DataFrame
                            numeric_values = []
                            for val in df[col]:
                                if pd.notna(val) and isinstance(val, str) and val.startswith('$'):
                                    # Remove $ and commas, convert to float
                                    clean_val = val.replace('$', '').replace(',', '')
                                    try:
                                        numeric_values.append(float(clean_val))
                                    except:
                                        pass
                                elif pd.notna(val) and isinstance(val, (int, float)):
                                    numeric_values.append(val)
                            total_value = sum(numeric_values) if numeric_values else 0
                            totals_row[col] = f"${total_value:,.2f}"
                        except:
                            totals_row[col] = "$0.00"
                else:
                    totals_row[col] = ""
            
            df_totals = pd.DataFrame([totals_row])
            df_final = pd.concat([df, df_totals], ignore_index=True)
            
            # Export to Excel
            with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                df_final.to_excel(writer, sheet_name='Subsections Summary', index=False)
                
                # Apply formatting
                workbook = writer.book
                worksheet = writer.sheets['Subsections Summary']
                
                # Auto-adjust column widths
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
                
                # Style header row
                from openpyxl.styles import Font, PatternFill, Alignment
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                header_alignment = Alignment(horizontal="center", vertical="center")
                
                for cell in worksheet[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                
                # Style totals row
                totals_row_num = len(df_final) + 1
                totals_font = Font(bold=True, color="FFFFFF")
                totals_fill = PatternFill(start_color="C0504D", end_color="C0504D", fill_type="solid")
                totals_alignment = Alignment(horizontal="center", vertical="center")
                
                for col in range(1, worksheet.max_column + 1):
                    cell = worksheet.cell(row=totals_row_num, column=col)
                    cell.font = totals_font
                    cell.fill = totals_fill
                    cell.alignment = totals_alignment
            
            logger.info(f"Generated subsections summary Excel: {filepath}")
            return str(filepath)
            
        except Exception as e:
            logger.error(f"Error generating subsections summary Excel: {str(e)}")
            raise

    def export_boq_items(self, items, grand_totals=None):
        """Export BOQ items to Excel with optional grand totals"""
        try:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"boq_items_{timestamp}.xlsx"
            filepath = self.exports_dir / filename
            
            if not items:
                raise ValueError("No data to export")
            
            # Define column order to match BOQ table order
            all_possible_headers = [
                'serial_number', 'structure', 'system', 'section_number', 'description', 'unit',
                'price', 'original_contract_quantity', 'total_contract_sum'
            ]
            
            # Add contract update quantity columns in order
            for key in items[0].keys():
                if key.startswith('updated_contract_quantity_'):
                    all_possible_headers.append(key)
            
            # Add contract update sum columns in order
            for key in items[0].keys():
                if key.startswith('updated_contract_sum_'):
                    all_possible_headers.append(key)
            
            all_possible_headers.extend([
                'estimated_quantity', 'quantity_submitted', 'internal_quantity',
                'approved_by_project_manager', 'approved_signed_quantity', 'total_estimate',
                'total_submitted', 'internal_total', 'total_approved_by_project_manager',
                'approved_signed_total', 'subsection', 'notes'
            ])
            
            # Only include headers that exist in the data
            ordered_headers = [h for h in all_possible_headers if h in items[0].keys()]
            
            # Create DataFrame with ordered columns
            df = pd.DataFrame(items)[ordered_headers]
            
            # Create a copy of original numeric data for totals calculation
            original_numeric_data = {}
            for col in df.columns:
                if df[col].dtype in ['float64', 'int64']:
                    original_numeric_data[col] = df[col].copy()
                elif col in ['total_contract_sum', 'total_estimate', 'total_submitted', 'internal_total', 'total_approved', 'approved_signed_total'] or col.startswith('total_updated_contract_sum_'):
                    # Force conversion to numeric for known total columns
                    try:
                        numeric_series = pd.to_numeric(df[col], errors='coerce')
                        original_numeric_data[col] = numeric_series.fillna(0)
                    except:
                        pass
            
            # Format only price and sum/total columns (not quantity columns)
            for col in df.columns:
                if col in df.columns and df[col].dtype in ['float64', 'int64']:
                    # Only apply $ formatting to price and sum/total columns, not quantity columns
                    if ('total' in col.lower() or 'sum' in col.lower() or 'price' in col.lower()) and 'quantity' not in col.lower():
                        df[col] = df[col].apply(lambda x: f"${x:,.2f}" if pd.notna(x) and isinstance(x, (int, float)) else "$0.00")
                    else:
                        # Format quantity columns without $ symbol
                        df[col] = df[col].apply(lambda x: f"{x:,.2f}" if pd.notna(x) and isinstance(x, (int, float)) and x != int(x) else str(int(x)) if pd.notna(x) and isinstance(x, (int, float)) else "0")
            
            # Define columns that should have grand totals
            total_columns = {
                'total_contract_sum',
                'total_estimate',
                'total_submitted',
                'internal_total',
                'total_approved',
                'approved_signed_total'
            }
            # Add updated contract sum columns
            for col in df.columns:
                if col.startswith('updated_contract_sum_'):
                    total_columns.add(col)
            
            # Calculate grand totals row using original numeric data - only for specified columns
            totals_row = {}
            for col in df.columns:
                if col == 'section_number':
                    totals_row[col] = "GRAND TOTAL"
                elif col in total_columns and col in original_numeric_data:
                    # Use provided grand totals if available, otherwise calculate from original numeric data
                    if grand_totals and col in grand_totals:
                        total_value = grand_totals[col]
                    else:
                        # Calculate totals from original numeric data
                        total_value = sum([val for val in original_numeric_data[col] if pd.notna(val)])
                    
                    # Apply currency formatting for all total columns
                    totals_row[col] = f"${total_value:,.2f}" if isinstance(total_value, (int, float)) else "$0.00"
                else:
                    totals_row[col] = ""
            
            df_totals = pd.DataFrame([totals_row])
            df_final = pd.concat([df, df_totals], ignore_index=True)
            
            # Export to Excel
            with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                df_final.to_excel(writer, sheet_name='BOQ Items', index=False)
                
                # Apply formatting
                workbook = writer.book
                worksheet = writer.sheets['BOQ Items']
                
                # Auto-adjust column widths
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
                
                # Style header row
                from openpyxl.styles import Font, PatternFill, Alignment
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                header_alignment = Alignment(horizontal="center", vertical="center")
                
                for cell in worksheet[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                
                # Style totals row
                totals_row_num = len(df_final) + 1
                totals_font = Font(bold=True, color="FFFFFF")
                totals_fill = PatternFill(start_color="C0504D", end_color="C0504D", fill_type="solid")
                totals_alignment = Alignment(horizontal="center", vertical="center")
                
                for col in range(1, worksheet.max_column + 1):
                    cell = worksheet.cell(row=totals_row_num, column=col)
                    cell.font = totals_font
                    cell.fill = totals_fill
                    cell.alignment = totals_alignment
            
            logger.info(f"Generated BOQ items Excel: {filepath}")
            return str(filepath)
            
        except Exception as e:
            logger.error(f"Error generating BOQ items Excel: {str(e)}")
            raise 